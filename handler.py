import csv
import os

import openpyxl


class Analysis():
    ''' Handles conversion of stow data aka data prep '''
    data_entries = []
    employees_completed = []

    key_employee_id = "Employee Id"
    key_employee = "Employee Name"
    key_units = "Units"
    key_uph = "UPH"

    def __init__(self, file):
        self.csv_file = csv.reader(file)
        column_headers = self.get_headers(file)

        # verify if file has columns/data needed
        if not self.has_required_columns(column_headers):
            raise Exception(
                "File does not have required columns of data to analyze correctly.")

        self.columns = {
            'employee_id': column_headers.index(Analysis.key_employee_id),
            'employee': column_headers.index(Analysis.key_employee),
            'units': column_headers.index(Analysis.key_units),
            'uph': column_headers.index(Analysis.key_uph)
        }

    def analyze(self, sort=True):
        ''' Performs analysis for each row and saves the output to a csv. '''
        for row in self.csv_file:
            self.update_row_data(row)

            if self.employee_id not in Analysis.employees_completed:
                Analysis.employees_completed.append(self.employee_id)
                self.append_data_entry()

            if sort:
                Analysis.sorted_by_rate(Analysis.data_entries)

    def append_data_entry(self):
        Analysis.data_entries.append(
            [
                self.employee,
                self.units,
                self.rate
            ]
        )

    def update_row_data(self, row):
        ''' Grabs and updates instance values with current row data '''
        if row[self.columns['uph']] != '0':
            self.employee_id = row[0]
            self.employee = row[self.columns['employee']]
            self.units = int(row[self.columns['units']])
            self.units_per_hour = float(row[self.columns['uph']])
            self.rate = Analysis.calculate_rate(self.units_per_hour)

    def get_headers(self, file):
        column_header = str(file.readline()).split(',')
        return column_header

    def has_required_columns(self, col):
        ''' Validates that the file has the necessary data columns '''
        return (
            Analysis.key_employee_id in col and
            Analysis.key_employee in col and
            Analysis.key_units in col and
            Analysis.key_uph in col
        )

    @staticmethod
    def calculate_rate(units_per_hour):
        return float('{:.1f}'.format(units_per_hour / 60))

    @staticmethod
    def sorted_by_rate(entries, increasing_order=False):
        if not increasing_order:
            entries.sort(key=lambda x: x[len(x)-1],
                         reverse=(increasing_order == False))
        else:
            entries.sort(key=lambda x: x[len(x)-1])
        return entries


class Excel():
    ''' Uses openpyxl to handle excel doc customizations '''

    @staticmethod
    def append_data_to_template(data, save_as=None):
        template_file = os.getcwd() + "\\Excel Template\\stow rate template.xlsx"
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active

        file_name = 'sample' if save_as is None else save_as

        for row, info in enumerate(data, start=1):
            print(row, info)
            ws['A{}'.format(row)] = info[0]
            ws['B{}'.format(row)] = info[1]
            ws['C{}'.format(row)] = info[2]

        wb.save(file_name + '.xlsx')
